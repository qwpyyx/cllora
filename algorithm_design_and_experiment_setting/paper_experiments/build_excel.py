#!/usr/bin/env python3
"""Build paper_results_master.xlsx — read 总数据, auto-generate 5 summary sheets."""
import os, sys, statistics
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

DIR = os.path.dirname(os.path.abspath(__file__))
CSV_PATH = os.path.join(DIR, "paper_results_master.csv")
XL_PATH = os.path.join(DIR, "paper_results_master.xlsx")

# ---- Styling helpers ----
HEADER_FONT = Font(name="Times New Roman", bold=True, size=10)
DATA_FONT = Font(name="Times New Roman", size=10)
TITLE_FONT = Font(name="Times New Roman", bold=True, size=12)
ITALIC_FONT = Font(name="Times New Roman", italic=True, size=9)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
HDR_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
OURS_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT; cell.alignment = CENTER; cell.border = THIN; cell.fill = HDR_FILL


def style_row(ws, row, ncols, is_ours=False):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = DATA_FONT; cell.alignment = CENTER; cell.border = THIN
        if is_ours:
            cell.fill = OURS_FILL


def auto_width(ws, ncols, w=14):
    for c in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = w


def cell_val(row, key):
    """Safely get a cell value, converting None to ''."""
    v = row.get(key)
    return v if v is not None else ""


# ============================================================
# Load CSV into list of dicts
# ============================================================
import csv

def load_csv():
    rows = []
    with open(CSV_PATH, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for r in reader:
            # Only convert metric value columns to float
            for k in ["primary_value", "secondary_value", "gen_len"]:
                try: r[k] = float(r[k])
                except: pass
            rows.append(r)
    return rows


def mean_std(values):
    """Return (mean, std) rounded to 2 dp. std=None if <2 values."""
    vals = [v for v in values if v is not None]
    if len(vals) < 2:
        return (round(vals[0], 2) if vals else None, None)
    m = statistics.mean(vals)
    s = statistics.stdev(vals)
    return (round(m, 2), round(s, 2))


def filter_rows(rows, **kwargs):
    """Filter rows by exact key=value matches (string comparison)."""
    out = []
    for r in rows:
        match = True
        for k, v in kwargs.items():
            if str(r.get(k, "")) != str(v):
                match = False
                break
        if match:
            out.append(r)
    return out


def get_method_seeds(rows, model, dataset, method, **extra):
    """Get primary_values for a method across seeds, sorted by seed."""
    sub = filter_rows(rows, model=model, dataset=dataset, method=method, **extra)
    # Sort by seed
    sub.sort(key=lambda r: int(r.get("seed", 0)))
    return [r.get("primary_value") for r in sub]


# ============================================================
# Build workbook
# ============================================================
def build():
    rows = load_csv()
    print(f"Loaded {len(rows)} rows from CSV")

    wb = Workbook()

    # ── Sheet 0: 总数据 ──
    ws0 = wb.active
    ws0.title = "总数据"
    # Write CSV headers
    with open(CSV_PATH, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        csv_headers = next(reader)
    for c, h in enumerate(csv_headers, 1):
        ws0.cell(row=1, column=c, value=h)
    style_header(ws0, 1, len(csv_headers))
    for i, r in enumerate(rows, 2):
        for c, h in enumerate(csv_headers, 1):
            ws0.cell(row=i, column=c, value=r.get(h, ""))
        style_row(ws0, i, len(csv_headers))
    auto_width(ws0, len(csv_headers), 16)

    # ── Sheet 1: Table 1 — Qwen3-14B NLP ──
    ws1 = wb.create_sheet("Table1_Qwen3_NLP")
    headers1 = ["Method", "GSM8K EM", "±Std", "Dolly ROUGE-L", "±Std", "Avg Rel Dense %", "Status"]
    for c, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=c, value=h)
    style_header(ws1, 1, len(headers1))

    METHODS_ORDER = [
        "Dense FedAvg-LoRA", "FLASC", "ComPEFT",
        "FLM-TopK", "FedComp", "FedSP-LoRA",
    ]

    # Pre-compute Dense reference for Avg Rel Dense (main_table only)
    dense_gsm8k = get_method_seeds(rows, "Qwen3-14B", "GSM8K", "Dense FedAvg-LoRA", alpha="10", data_type="main_table")
    dense_dolly = get_method_seeds(rows, "Qwen3-14B", "Dolly-15K", "Dense FedAvg-LoRA", data_type="main_table")
    dense_gsm8k_mean = statistics.mean(dense_gsm8k) if dense_gsm8k else 1
    dense_dolly_mean = statistics.mean(dense_dolly) if dense_dolly else 1

    for i, method in enumerate(METHODS_ORDER, 2):
        gsm = get_method_seeds(rows, "Qwen3-14B", "GSM8K", method, alpha="10", data_type="main_table")
        dolly = get_method_seeds(rows, "Qwen3-14B", "Dolly-15K", method, data_type="main_table")

        gsm_m, gsm_s = mean_std(gsm) if gsm else (None, None)
        dolly_m, dolly_s = mean_std(dolly) if dolly else (None, None)

        # Average Relative to Dense
        if gsm_m and dolly_m:
            avg_rel = round(((gsm_m / dense_gsm8k_mean) + (dolly_m / dense_dolly_mean)) / 2 * 100, 1)
        elif gsm_m:
            avg_rel = round(gsm_m / dense_gsm8k_mean * 100, 1)
        else:
            avg_rel = None

        gsm_str = f"{gsm_m}" if gsm_m else "TBD"
        gsm_std = f"±{gsm_s}" if gsm_s else "—"
        dolly_str = f"{dolly_m}" if dolly_m else "TBD"
        dolly_std = f"±{dolly_s}" if dolly_s else "—"
        rel_str = f"{avg_rel}" if avg_rel is not None else "TBD"

        status = "✅" if gsm_m and dolly_m else ("⚠️ partial" if gsm_m or dolly_m else "⏳ Pending")

        vals = [method, gsm_str, gsm_std, dolly_str, dolly_std, rel_str, status]
        for c, v in enumerate(vals, 1):
            ws1.cell(row=i, column=c, value=v)
        style_row(ws1, i, len(headers1), is_ours=("FedSP-LoRA" in method))

    n1 = len(headers1)
    ws1.merge_cells(start_row=len(METHODS_ORDER) + 3, start_column=1, end_row=len(METHODS_ORDER) + 3, end_column=n1)
    ws1.cell(row=len(METHODS_ORDER) + 3, column=1,
             value="GSM8K: 12.5% budget, alpha=10 quantity. Dolly: 12.5% budget, alpha=0.5 label-skew. All 3 seeds.").font = ITALIC_FONT
    auto_width(ws1, n1, 18)

    # ── Sheet 2: Table 2 — Llama-3.1-8B ──
    ws2 = wb.create_sheet("Table2_Llama3_NLP")
    for c, h in enumerate(headers1, 1):
        ws2.cell(row=1, column=c, value=h)
    style_header(ws2, 1, len(headers1))

    for i, method in enumerate(METHODS_ORDER, 2):
        gsm = get_method_seeds(rows, "Llama-3.1-8B", "GSM8K", method, alpha="10")
        dolly = get_method_seeds(rows, "Llama-3.1-8B", "Dolly-15K", method)

        gsm_m, gsm_s = mean_std(gsm) if gsm else (None, None)
        dolly_m, dolly_s = mean_std(dolly) if dolly else (None, None)

        gsm_str = f"{gsm_m}" if gsm_m else "TBD"
        gsm_std = f"±{gsm_s}" if gsm_s else "—"
        dolly_str = f"{dolly_m}" if dolly_m else "TBD"
        dolly_std = f"±{dolly_s}" if dolly_s else "—"

        has_data = gsm_m or dolly_m
        status = "✅" if gsm_m and dolly_m else ("⚠️ partial" if has_data else "⏳ Pending")

        vals = [method, gsm_str, gsm_std, dolly_str, dolly_std, "TBD", status]
        for c, v in enumerate(vals, 1):
            ws2.cell(row=i, column=c, value=v)
        style_row(ws2, i, len(headers1), is_ours=("FedSP-LoRA" in method))

    auto_width(ws2, n1, 18)

    # ── Sheet 3: Table 3 — Gemma + MoE ──
    ws3 = wb.create_sheet("Table3_Robustness")
    headers3 = ["Method", "Gemma-2B EM", "Qwen3-30B-A3B EM", "Avg Rel Dense %", "Status"]
    for c, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=c, value=h)
    style_header(ws3, 1, len(headers3))

    MODELS_MINI = ["Gemma-2B", "Qwen3-30B-A3B"]
    for i, method in enumerate(METHODS_ORDER, 2):
        vals_list = []
        for model in MODELS_MINI:
            vals = get_method_seeds(rows, model, "GSM8K", method)
            m, s = mean_std(vals) if vals else (None, None)
            vals_list.append(f"{m}" if m else "TBD")

        has_any = any(v != "TBD" for v in vals_list)
        status = "✅" if all(v != "TBD" for v in vals_list) else ("⚠️ partial" if has_any else "⏳ Pending")
        row_vals = [method] + vals_list + ["TBD", status]
        for c, v in enumerate(row_vals, 1):
            ws3.cell(row=i, column=c, value=v)
        style_row(ws3, i, len(headers3), is_ours=("FedSP-LoRA" in method))

    n3 = len(headers3)
    auto_width(ws3, n3, 18)

    # ── Sheet 4: Figure 2 — Budget Curve + VLM ──
    ws4 = wb.create_sheet("Fig2_BudgetCurve_VLM")

    # Part A: GSM8K Budget Curve
    ws4.merge_cells("A1:F1")
    ws4.cell(row=1, column=1, value="Part A: GSM8K Budget Curve (Qwen3-14B, seed28, alpha=10)").font = TITLE_FONT

    budgets_sorted = ["440", "880", "1760", "0"]
    budget_labels = {"440": "3.1%", "880": "6.25%", "1760": "12.5%", "0": "100% (Dense)"}

    h4a = ["Method"] + [f"b={b} ({budget_labels[b]})" for b in budgets_sorted]
    for c, h in enumerate(h4a, 1):
        ws4.cell(row=2, column=c, value=h)
    style_header(ws4, 2, len(h4a))

    BC_METHODS = ["FedSP-LoRA", "FLASC"]
    for i, method in enumerate(BC_METHODS, 3):
        vals = [method]
        for b in budgets_sorted:
            sub = filter_rows(rows, model="Qwen3-14B", dataset="GSM8K", method=method,
                              budget=b, alpha="10", seed="28", data_type="budget_curve")
            v = sub[0].get("primary_value") if sub else None
            vals.append(f"{v:.2f}" if v else "TBD")
        for c, v in enumerate(vals, 1):
            ws4.cell(row=i, column=c, value=v)
        style_row(ws4, i, len(h4a), is_ours=("FedSP-LoRA" in method))

    # Part B: VLM placeholder
    ws4.merge_cells("A7:F7")
    ws4.cell(row=7, column=1, value="Part B: VLM Budget Curve (Fed-SLAKE VQA, pending)").font = TITLE_FONT
    h4b = ["Method", "6.25%", "12.5%", "25%", "100%", "Status"]
    for c, h in enumerate(h4b, 1):
        ws4.cell(row=8, column=c, value=h)
    style_header(ws4, 8, len(h4b))
    VLM_METHODS = ["FedSP-LoRA", "FLASC", "FLM-TopK"]
    for i, method in enumerate(VLM_METHODS, 9):
        vals = [method] + ["TBD"] * 4 + ["⏳ Pending"]
        for c, v in enumerate(vals, 1):
            ws4.cell(row=i, column=c, value=v)
        style_row(ws4, i, len(h4b), is_ours=("FedSP-LoRA" in method))

    auto_width(ws4, 6, 20)

    # ── Sheet 5: Figure 3 — Ablation ──
    ws5 = wb.create_sheet("Fig3_Ablation")

    ws5.merge_cells("A1:E1")
    ws5.cell(row=1, column=1,
             value="Part A: Completed Ablation (Qwen3-14B, GSM8K, seed28, 12.5%, alpha=10)").font = TITLE_FONT

    h5a = ["Variant", "EM", "Δ vs Full", "P1", "P2"]
    for c, h in enumerate(h5a, 1):
        ws5.cell(row=2, column=c, value=h)
    style_header(ws5, 2, len(h5a))

    full_em = None
    sub = filter_rows(rows, model="Qwen3-14B", dataset="GSM8K", method="FedSP-LoRA",
                       data_type="main_table", seed="28", alpha="10", budget="1760")
    if sub: full_em = sub[0].get("primary_value")

    AB_METHODS = [
        ("FedSP-LoRA", "FedSP-LoRA"),
        ("FedSP-LoRA w/o depth calibration", "FedSP-LoRA w/o depth calibration"),
        ("FedSP-LoRA w/o P2", "FedSP-LoRA w/o P2"),
        ("FedSP-LoRA w/ A/B-pair atom", "FedSP-LoRA w/ A/B-pair atom"),
    ]

    def get_em(method_name):
        sub = filter_rows(rows, model="Qwen3-14B", dataset="GSM8K",
                           method=method_name, seed="28", alpha="10", budget="1760", data_type="ablation")
        return sub[0].get("primary_value") if sub else None

    for i, (display_name, lookup_name) in enumerate(AB_METHODS, 3):
        em = get_em(lookup_name)
        delta = f"{(em - full_em):+.2f}" if em and full_em else "—"
        em_str = f"{em:.2f}" if em else "TBD"

        # P1/P2 flags
        if display_name == "FedSP-LoRA":
            p1, p2 = "✓ depth_rank", "✓ gap=1.0"
        elif "w/o depth" in display_name:
            p1, p2 = "✓ (raw)", "✓"
        elif "w/o P2" in display_name:
            p1, p2 = "✓", "✗"
        elif "A/B-pair" in display_name:
            p1, p2 = "✓ (ab_pair)", "✓"
        else:
            p1, p2 = "—", "—"

        vals = [display_name, em_str, delta, p1, p2]
        for c, v in enumerate(vals, 1):
            ws5.cell(row=i, column=c, value=v)
        style_row(ws5, i, len(h5a), is_ours=(display_name == "FedSP-LoRA"))

    n5 = len(h5a)

    # Part B
    ws5.merge_cells(start_row=8, start_column=1, end_row=8, end_column=n5)
    ws5.cell(row=8, column=1,
             value="Part B: Planned Final Ablation (P2-sensitive setting: FinGPT/Dolly, label-skew, pending)").font = TITLE_FONT
    for c, h in enumerate(["Variant", "P1", "P2", "Expected Behavior", "Status"], 1):
        ws5.cell(row=9, column=c, value=h)
    style_header(ws5, 9, 5)

    planned = [
        ("AB-Effective", "✗", "✗", "Magnitude anchor", "⏳"),
        ("FedSP-LoRA w/o P2", "✓", "✗", "P1-only, shows P1 signal-noise value", "⏳"),
        ("FedSP-LoRA w/o P1 depth", "✓(raw)", "✓", "P1+P2 without calibration", "⏳"),
        ("FedSP-LoRA", "✓", "✓", "Full method", "⏳"),
    ]
    for i, row in enumerate(planned, 10):
        for c, v in enumerate(row, 1):
            ws5.cell(row=i, column=c, value=v)
        style_row(ws5, i, 5, is_ours=("FedSP-LoRA" == row[0]))

    auto_width(ws5, 5, 22)

    # ── Save ──
    wb.save(XL_PATH)
    print(f"Saved to {XL_PATH}")


if __name__ == "__main__":
    build()
